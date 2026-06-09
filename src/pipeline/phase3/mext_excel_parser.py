"""
文部科学省 Excel 解析器
数据源: https://www.mext.go.jp/a_menu/koutou/ichiran/mext_00038.html
令和6年度全国大学一覧 (Excel)

每个 Sheet = 一所大学
列结构（典型）:
  - 学部名 / 学科名 / 入学定员
  - 研究科名 / 専攻名 / 入学定员
  - 所在地（都道府県）

解析结果 → 批量导入 university_units 表
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ── 识别"研究科"关键词 ──────────────────────────────────────────
GRADUATE_KEYWORDS = ["研究科", "大学院"]
UNDERGRADUATE_KEYWORDS = ["学部", "学院", "学群", "学類"]

# 都道府県一覧（用于从单元格文本中提取）
PREFECTURES = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
    "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
    "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
    "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
    "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
    "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
    "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
]


def extract_prefecture(text: str) -> Optional[str]:
    """从地址字符串中提取都道府県名"""
    if not text:
        return None
    for pref in PREFECTURES:
        if pref in text:
            return pref
    return None


def is_graduate_unit(name: str) -> bool:
    return any(kw in name for kw in GRADUATE_KEYWORDS)


def is_undergraduate_unit(name: str) -> bool:
    return any(kw in name for kw in UNDERGRADUATE_KEYWORDS)


def clean_cell(cell_value) -> Optional[str]:
    """清洗单元格值：去除空白、零宽字符等"""
    if cell_value is None:
        return None
    text = str(cell_value).strip()
    # 去除零宽空格等不可见字符
    text = re.sub(r"[\u200b\u200c\u200d\ufeff\u3000]", "", text)
    # 合并连续空白
    text = re.sub(r"\s+", " ", text).strip()
    return text if text else None


def _is_noise_unit_name(text: str) -> bool:
    """过滤 Excel 中的合计、注释、纯数字等非学部/研究科名称。"""
    if not text:
        return True
    normalized = text.strip()
    if normalized in {"計", "合計", "小計", "総計", "－", "-", "名称"}:
        return True
    if normalized in {"学部", "学科", "研究科", "専攻", "都道府県", "市区町村"}:
        return True
    if re.fullmatch(r"[\d,\.]+", normalized):
        return True
    if normalized.startswith(("注", "※")):
        return True
    return False


def clean_university_name(name: str) -> str:
    """统一大学名：去掉设置者前缀、英文括号等。"""
    if not name:
        return name
    name = re.sub(r"[（(][A-Za-z0-9\s\-,.&]+[）)]", "", name).strip()
    name = re.sub(r"^(国立|公立|私立)\s+", "", name).strip()
    return name


class MextExcelParser:
    """
    解析文部科学省全国大学一覧 Excel 文件。

    使用方法:
        parser = MextExcelParser("R06_daigaku_ichiran.xlsx")
        records = parser.parse()
        # records: List[UniversityUnitRecord]
    """

    def __init__(self, filepath: str | Path):
        self.filepath = Path(filepath)
        self._workbook = None

    def _load(self):
        if self._workbook is None:
            logger.info("加载 Excel: %s", self.filepath)
            self._workbook = openpyxl.load_workbook(
                self.filepath, read_only=True, data_only=True
            )

    def parse(self) -> list[dict]:
        """
        解析所有 Sheet，返回 university_units 记录列表。

        返回格式:
            [
                {
                    "university_name": "北海道大学",
                    "unit_type": "学部",
                    "unit_name": "文学部",
                    "sub_unit_name": "人文科学科",
                    "prefecture": "北海道",
                },
                ...
            ]
        """
        self._load()
        all_records: list[dict] = []

        for sheet_name in self._workbook.sheetnames:
            sheet = self._workbook[sheet_name]
            try:
                records = self._parse_sheet(sheet, sheet_name)
                all_records.extend(records)
                logger.info("Sheet '%s': %d 件", sheet_name, len(records))
            except Exception as e:
                logger.warning("Sheet '%s' 解析失败: %s", sheet_name, e)

        logger.info("合計 %d 件解析完了", len(all_records))
        return all_records

    def _parse_sheet(self, sheet: Worksheet, sheet_name: str) -> list[dict]:
        """
        解析单个 Sheet（一所大学）。

        文部科学省 Excel 的列结构因年度而略有不同，
        此处实现自适应解析：先扫描表头，再按列名提取。
        """
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        # ── Step 1: 提取大学名 ──────────────────────────────
        university_name = self._extract_university_name(rows, sheet_name)

        # ── Step 2: 提取都道府県 ────────────────────────────
        prefecture = self._extract_prefecture_from_rows(rows)

        # ── Step 3: 优先解析文科省标准「学部」「研究科」区块 ─────────
        # 该 Excel 前部还有「学部・研究科所在地」等区域，不能当作 ground truth。
        section_records = self._parse_mext_unit_sections(rows, university_name, prefecture)
        if section_records:
            return section_records

        # ── Step 4: 定位表头行，识别列索引（fallback）────────────
        header_row_idx, col_map = self._find_header_and_columns(rows)
        if header_row_idx is None:
            logger.warning("Sheet '%s': 未找到表头行", sheet_name)
            return self._fallback_parse(rows, university_name, prefecture)

        # ── Step 5: 按行提取 unit 数据（fallback）───────────────
        records = self._extract_units(
            rows, header_row_idx, col_map,
            university_name, prefecture
        )
        return records

    def _extract_university_name(self, rows: list, sheet_name: str) -> str:
        """
        从 Sheet 前几行提取大学名。
        通常第1~3行中含有完整大学名（如 "○○大学"）。
        """
        for row in rows[:5]:
            for cell in row:
                val = clean_cell(cell)
                if val and ("大学" in val or "短期大学" in val):
                    # 去掉括号内的英文等杂项
                    name = re.sub(r"[（(].*?[）)]", "", val).strip()
                    if len(name) >= 3:  # 至少3个字符
                        return clean_university_name(name)
        # fallback: 用 Sheet 名
        return clean_university_name(sheet_name)

    def _extract_prefecture_from_rows(self, rows: list) -> Optional[str]:
        """扫描前20行，寻找都道府県信息"""
        for row in rows[:20]:
            for cell in row:
                val = clean_cell(cell)
                if val:
                    pref = extract_prefecture(val)
                    if pref:
                        return pref
        return None

    def _parse_mext_unit_sections(
        self,
        rows: list,
        university_name: str,
        prefecture: Optional[str],
    ) -> list[dict]:
        """
        解析文科省大学一覧 Excel 的标准主区块：
          - 「学部」区块：表头为 学部 / 学科 / 都道府県 / 市区町村 ...
          - 「研究科」区块：表头为 研究科 / 専攻 / 都道府県 / 市区町村 ...

        这样可以避开文件前部的「学部・研究科所在地」和后部的沿革/通信教育等区域。
        """
        records: list[dict] = []

        def cell(row: tuple, idx: int) -> Optional[str]:
            return clean_cell(row[idx]) if idx < len(row) else None

        def is_undergrad_header(row: tuple) -> bool:
            return cell(row, 1) == "学部" and cell(row, 3) == "学科" and cell(row, 6) == "都道府県"

        def is_grad_header(row: tuple) -> bool:
            return cell(row, 1) == "研究科" and cell(row, 3) == "専攻" and cell(row, 6) == "都道府県"

        def parse_block(start_idx: int, unit_type: str) -> None:
            started = False
            for row in rows[start_idx + 1:]:
                unit_name = cell(row, 1)
                sub_unit_name = cell(row, 3)
                row_prefecture = cell(row, 6) or prefecture

                if unit_name in {"通信教育部（学部）", "通信教育部（研究科）", "専攻科", "別科"}:
                    break
                if unit_name and any(marker in unit_name for marker in ["共同実施制度", "連係課程", "沿革", "連合研究科"]):
                    break

                if not unit_name and not sub_unit_name:
                    if started:
                        break
                    continue

                # 研究科表头有两行，第二行是「入学定員 / 修業年限 ...」，要跳过。
                if _is_noise_unit_name(unit_name or ""):
                    continue
                if sub_unit_name and _is_noise_unit_name(sub_unit_name):
                    sub_unit_name = None

                # 只接受形态合理的主组织名称，避免把定員等行混入。
                if unit_type == "学部" and "学部" not in unit_name:
                    continue
                if unit_type == "研究科" and not any(suffix in unit_name for suffix in ["研究科", "学院", "教育部"]):
                    continue

                records.append({
                    "university_name": clean_university_name(university_name),
                    "unit_type": unit_type,
                    "unit_name": unit_name,
                    "sub_unit_name": sub_unit_name,
                    "prefecture": row_prefecture or prefecture,
                })
                started = True

        undergrad_header_idx: Optional[int] = None
        grad_header_idx: Optional[int] = None

        for idx, row in enumerate(rows):
            if undergrad_header_idx is None and is_undergrad_header(row):
                undergrad_header_idx = idx
            elif grad_header_idx is None and is_grad_header(row):
                grad_header_idx = idx

            if undergrad_header_idx is not None and grad_header_idx is not None:
                break

        if undergrad_header_idx is not None:
            parse_block(undergrad_header_idx, "学部")
        if grad_header_idx is not None:
            parse_block(grad_header_idx, "研究科")

        # 去重
        seen = set()
        unique_records: list[dict] = []
        for r in records:
            key = (r["university_name"], r["unit_type"], r["unit_name"], r["sub_unit_name"])
            if key not in seen:
                seen.add(key)
                unique_records.append(r)

        return unique_records

    def _find_header_and_columns(
        self, rows: list
    ) -> tuple[Optional[int], dict[str, int]]:
        """
        扫描行，找到包含"学部"/"研究科"/"学科"/"専攻"等关键词的表头行。
        返回 (表头行索引, {列名: 列索引}) 。
        """
        header_keywords = {
            "大学名": ["大学名", "学校名", "大学名称", "学校名称"],
            "学部名": ["学部名", "学部・学科名", "学部等", "学部"],
            "学科名": ["学科名", "学科"],
            "研究科名": ["研究科名", "研究科", "大学院研究科"],
            "専攻名": ["専攻名", "専攻"],
            "入学定員": ["入学定員", "定員"],
            "所在地": ["所在地", "主たる所在地", "所 在 地", "都道府県"],
        }

        for row_idx, row in enumerate(rows):
            cells = [clean_cell(c) for c in row]
            matched: dict[str, int] = {}
            for canonical, aliases in header_keywords.items():
                for alias in aliases:
                    for col_idx, cell_val in enumerate(cells):
                        if cell_val and alias in cell_val:
                            matched[canonical] = col_idx
                            break
                    if canonical in matched:
                        break

            # 如果本行至少匹配到 2 个关键词，认为是表头
            if len(matched) >= 2:
                return row_idx, matched

        return None, {}

    def _extract_units(
        self,
        rows: list,
        header_row_idx: int,
        col_map: dict[str, int],
        university_name: str,
        prefecture: Optional[str],
    ) -> list[dict]:
        """按表头列索引提取 unit 记录"""
        records: list[dict] = []

        current_university_name: str = university_name
        current_prefecture: Optional[str] = prefecture
        current_unit_type: Optional[str] = None
        current_unit_name: Optional[str] = None

        for row in rows[header_row_idx + 1:]:
            cells = [clean_cell(c) for c in row]
            if not any(cells):
                continue

            # ── 多大学同表格式：大学名列通常使用合并单元格，空值沿用上一行 ──
            if "大学名" in col_map:
                candidate = cells[col_map["大学名"]] if col_map["大学名"] < len(cells) else None
                if candidate and "計" not in candidate and "合計" not in candidate:
                    candidate = re.sub(r"[（(].*?[）)]", "", candidate).strip()
                    if "大学" in candidate and len(candidate) >= 3:
                        current_university_name = candidate
                        current_unit_type = None
                        current_unit_name = None

            # ── 所在地（可能细化都道府県）──────────────────
            if "所在地" in col_map:
                addr = cells[col_map["所在地"]] if col_map["所在地"] < len(cells) else None
                if addr:
                    detected = extract_prefecture(addr)
                    if detected:
                        current_prefecture = detected

            # ── 尝试读取学部名 / 研究科名 ──────────────────
            unit_name_val = None
            unit_type_val = None

            if "学部名" in col_map:
                val = cells[col_map["学部名"]] if col_map["学部名"] < len(cells) else None
                if val and not _is_noise_unit_name(val):
                    unit_name_val = val
                    unit_type_val = "研究科" if is_graduate_unit(val) else "学部"

            if "研究科名" in col_map and not unit_name_val:
                val = cells[col_map["研究科名"]] if col_map["研究科名"] < len(cells) else None
                if val and not _is_noise_unit_name(val):
                    unit_name_val = val
                    unit_type_val = "研究科"

            if unit_name_val:
                current_unit_type = unit_type_val
                current_unit_name = unit_name_val

            # ── 尝试读取学科名 / 専攻名 ────────────────────
            sub_unit_val = None
            if current_unit_type == "学部" and "学科名" in col_map:
                sub_unit_val = cells[col_map["学科名"]] if col_map["学科名"] < len(cells) else None
            elif current_unit_type == "研究科" and "専攻名" in col_map:
                sub_unit_val = cells[col_map["専攻名"]] if col_map["専攻名"] < len(cells) else None

            if sub_unit_val and _is_noise_unit_name(sub_unit_val):
                sub_unit_val = None

            # ── 仅在有 unit_name 时记录 ────────────────────
            if current_university_name and current_unit_name and current_unit_type:
                # 有 sub_unit 则记录 sub_unit 行，否则只记录 unit 行（避免重复）
                if sub_unit_val:
                    records.append({
                        "university_name": current_university_name,
                        "unit_type": current_unit_type,
                        "unit_name": current_unit_name,
                        "sub_unit_name": sub_unit_val,
                        "prefecture": current_prefecture,
                    })
                elif unit_name_val:
                    # unit 本身变化了，记录一条"仅 unit 级"记录
                    records.append({
                        "university_name": current_university_name,
                        "unit_type": current_unit_type,
                        "unit_name": current_unit_name,
                        "sub_unit_name": None,
                        "prefecture": current_prefecture,
                    })

        # 去重
        seen = set()
        unique_records = []
        for r in records:
            key = (r["university_name"], r["unit_type"], r["unit_name"], r["sub_unit_name"])
            if key not in seen:
                seen.add(key)
                unique_records.append(r)

        return unique_records

    def _fallback_parse(
        self, rows: list, university_name: str, prefecture: Optional[str]
    ) -> list[dict]:
        """
        表头识别失败时的 fallback：
        扫描所有单元格，寻找包含 "学部" / "研究科" 的文本。
        精度较低，仅保底用。
        """
        records: list[dict] = []
        for row in rows:
            for cell in row:
                val = clean_cell(cell)
                if not val or len(val) < 3 or _is_noise_unit_name(val):
                    continue
                if is_undergraduate_unit(val) and "学部" in val:
                    records.append({
                        "university_name": university_name,
                        "unit_type": "学部",
                        "unit_name": val,
                        "sub_unit_name": None,
                        "prefecture": prefecture,
                    })
                elif is_graduate_unit(val):
                    records.append({
                        "university_name": university_name,
                        "unit_type": "研究科",
                        "unit_name": val,
                        "sub_unit_name": None,
                        "prefecture": prefecture,
                    })

        # 去重
        seen = set()
        unique: list[dict] = []
        for r in records:
            key = (r["unit_type"], r["unit_name"])
            if key not in seen:
                seen.add(key)
                unique.append(r)
        return unique
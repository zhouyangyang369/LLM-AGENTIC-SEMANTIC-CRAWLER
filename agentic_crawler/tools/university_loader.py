"""
大学データ読み込み — Excel から官網 URL と sitemap URL をマージして返す。
"""

import logging
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from config import OFFICIAL_EXCEL, SITEMAP_EXCEL, SITEMAP_SHEETS

logger = logging.getLogger(__name__)


@dataclass
class UniversityInfo:
    name: str
    official_url: str = ""
    sitemap_url: str = ""
    university_type: str = "national"  # national / public / private

    @property
    def domain(self) -> str:
        from urllib.parse import urlparse
        parsed = urlparse(self.official_url or self.sitemap_url)
        return parsed.netloc.lstrip("www.")


def _read_sheet(wb: openpyxl.Workbook, sheet_name: str) -> dict[str, str]:
    """シートから {name: url} の dict を返す（ヘッダー行スキップ）"""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    result: dict[str, str] = {}
    for row in rows[1:]:  # 先頭ヘッダーをスキップ
        if len(row) >= 3 and row[1] and row[2]:
            name = str(row[1]).strip()
            url = str(row[2]).strip()
            result[name] = url
    return result


def load_universities(types: list[str] | None = None) -> list[UniversityInfo]:
    """
    指定タイプの大学一覧を読み込む。
    types: ["national", "public", "private"] のサブセット。None なら national のみ。
    """
    if types is None:
        types = ["national"]

    # ── 官網 URL 読み込み ───────────────────────────────
    official_map: dict[str, str] = {}
    for utype in types:
        excel_path = OFFICIAL_EXCEL.get(utype)
        if excel_path and Path(excel_path).exists():
            wb = openpyxl.load_workbook(excel_path)
            data = _read_sheet(wb, wb.sheetnames[0])
            for name, url in data.items():
                official_map[name] = url
        else:
            logger.warning(f"Official Excel not found for type '{utype}': {excel_path}")

    # ── Sitemap URL 読み込み ────────────────────────────
    sitemap_map: dict[str, str] = {}
    if Path(SITEMAP_EXCEL).exists():
        wb_sitemap = openpyxl.load_workbook(SITEMAP_EXCEL)
        for sheet_name, utype in SITEMAP_SHEETS.items():
            if utype not in types:
                continue
            data = _read_sheet(wb_sitemap, sheet_name)
            sitemap_map.update(data)
    else:
        logger.warning(f"Sitemap Excel not found: {SITEMAP_EXCEL}")

    # ── マージ ─────────────────────────────────────────
    # official_map を基準にマージ（sitemap にしかない学校も追加）
    all_names = set(official_map.keys()) | set(sitemap_map.keys())
    universities: list[UniversityInfo] = []
    for name in sorted(all_names):
        utype = "national"  # TODO: 複数タイプ対応時に拡張
        uni = UniversityInfo(
            name=name,
            official_url=official_map.get(name, ""),
            sitemap_url=sitemap_map.get(name, ""),
            university_type=utype,
        )
        universities.append(uni)

    logger.info(f"Loaded {len(universities)} universities (types={types})")
    return universities


def get_university(name: str, types: list[str] | None = None) -> UniversityInfo | None:
    """名前で一件取得"""
    for uni in load_universities(types):
        if uni.name == name:
            return uni
    return None

import os
import openpyxl

for fname in os.listdir('university_excel'):
    path = f'university_excel/{fname}'
    print('\n' + '='*70)
    print(f'FILE: {fname}')
    print('='*70)
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet in wb.worksheets:
        print(f'\n--- Sheet: {sheet.title} | Dims: {sheet.dimensions} | Rows: {sheet.max_row} | Cols: {sheet.max_column} ---')
        rows = list(sheet.iter_rows(values_only=True))
        for i, r in enumerate(rows[:6]):
            print(f'  Row{i}: {r}')
        if len(rows) > 6:
            print(f'  ... (total {len(rows)} rows)')